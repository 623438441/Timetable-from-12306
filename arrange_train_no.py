import re

import openpyxl
import pandas as pd

# coding: utf-8

date = input('YYYYMMDD')
with open('train_info.txt', 'r') as r:
    lines = r.readlines()
with open('train_info.txt', 'w') as w:
    for l in lines:
        if f'"start_train_date":"{date}"' in l:
            w.write(l)

# 读取存在空行的文件
fr = open('train_info.txt', 'r')
# 输出去掉空行的文件
fd = open('temp.txt', 'w')
for text in fr.readlines():
    if text.split():
        fd.write(text)
fd.close()
fr.close()
data = open('temp.txt', 'r')
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "train_info"
for lines in range(0, len(open("temp.txt").readlines())):
    line = data.readline()
    if line != '\n':
        # worksheet.cell(lines + 1, 1, str(re.findall(r'\"station_train_code\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 2, str(re.findall(r'\"train_no\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 3, str(re.findall(r'\"start_station_name\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 4, str(re.findall(r'\"start_station_telecode\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 5, str(re.findall(r'\"end_station_name\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 6, str(re.findall(r'\"end_station_telecode\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 7, str(re.findall(r'\"train_type_name\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 8, str(re.findall(r'\"train_class_code\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 9, str(re.findall(r'\"train_class_name\"\:\"([\s\S]+?)\"', line))[2:-2])
        worksheet.cell(lines + 1, 10, str(re.findall(r'\"seat_types\"\:\"([\s\S]+?)\"', line))[2:-2])
        # bureau_code = str(re.findall(r'\"bureau_code\"\:\"([\s\S]+?)\"', line))[2:-2]
        # if bureau_code == 'P':
        #     worksheet.cell(lines + 1, 11, '中国铁路北京局集团有限公司')
        # elif bureau_code == 'B':
        #     worksheet.cell(lines + 1, 11, '中国铁路哈尔滨局集团有限公司')
        # elif bureau_code == 'C':
        #     worksheet.cell(lines + 1, 11, '中国铁路呼和浩特局集团有限公司')
        # elif bureau_code == 'F':
        #     worksheet.cell(lines + 1, 11, '中国铁路郑州局集团有限公司')
        # elif bureau_code == 'G':
        #     worksheet.cell(lines + 1, 11, '中国铁路南昌局集团有限公司')
        # elif bureau_code == 'H':
        #     worksheet.cell(lines + 1, 11, '中国铁路上海局集团有限公司')
        # elif bureau_code == 'J':
        #     worksheet.cell(lines + 1, 11, '中国铁路兰州局集团有限公司')
        # elif bureau_code == 'K':
        #     worksheet.cell(lines + 1, 11, '中国铁路济南局集团有限公司')
        # elif bureau_code == 'M':
        #     worksheet.cell(lines + 1, 11, '中国铁路昆明局集团有限公司')
        # elif bureau_code == 'N':
        #     worksheet.cell(lines + 1, 11, '中国铁路武汉局集团有限公司')
        # elif bureau_code == 'O':
        #     worksheet.cell(lines + 1, 11, '中国铁路青藏集团有限公司')
        # elif bureau_code == 'Q':
        #     worksheet.cell(lines + 1, 11, '中国铁路广州局集团有限公司')
        # elif bureau_code == 'R':
        #     worksheet.cell(lines + 1, 11, '中国铁路乌鲁木齐局集团有限公司')
        # elif bureau_code == 'T':
        #     worksheet.cell(lines + 1, 11, '中国铁路沈阳局集团有限公司')
        # elif bureau_code == 'U':
        #     worksheet.cell(lines + 1, 11, '广东珠三角城际轨道交通有限公司')
        # elif bureau_code == 'V':
        #     worksheet.cell(lines + 1, 11, '中国铁路太原局集团有限公司')
        # elif bureau_code == 'W':
        #     worksheet.cell(lines + 1, 11, '中国铁路成都局集团有限公司')
        # elif bureau_code == 'X':
        #     worksheet.cell(lines + 1, 11, '港铁')
        # elif bureau_code == 'Y':
        #     worksheet.cell(lines + 1, 11, '中国铁路西安局集团有限公司')
        # elif bureau_code == 'Z':
        #     worksheet.cell(lines + 1, 11, '中国铁路南宁局集团有限公司')
        service_type = str(re.findall(r'\"service_type\"\:\"([\s\S]+?)\"', line))[2:-2]
        if service_type == '0':
            worksheet.cell(lines + 1, 11, '无空调')
        else:
            worksheet.cell(lines + 1, 11, '有空调')
        worksheet.cell(lines + 1, 12, str(re.findall(r'\"start_train_date\"\:\"([\s\S]+?)\"', line))[2:-2])
workbook.save(filename="train_info.xlsx")
xlsx = pd.read_excel('train_info.xlsx', index_col=0)
xlsx.to_csv('train_info.csv', encoding='utf-8')
csv = pd.read_csv('train_info.csv', low_memory=False, error_bad_lines=False, header=None)  # 读取csv中的数据
df = pd.DataFrame(csv)
f = df.drop_duplicates(subset=1)  # 去重
f.to_csv('trains.csv', index=None)  # 写到一个新的文件
new_csv = pd.read_csv('trains.csv', encoding='utf-8', header=0,
                      names=['TrainNumber', 'TrainCode', 'StartStation', 'StartTelecode', 'EndStation', 'EndTelecode',
                             'TrainType', 'TypeCode', 'TrainClass', 'SeatCode', 'Service', 'Date'])
new_csv.to_excel('train.xlsx', sheet_name='trains')
