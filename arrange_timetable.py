import datetime
import re

import openpyxl
import pandas as pd

data = open('origin_timetable.txt', 'r')
workbook = openpyxl.Workbook()
codelist = openpyxl.Workbook()
sheet = workbook.active
sheet2 = codelist.active
sheet.title = 'timetable'
sheet2.title = 'codelist'
line_gap = 2
for lines in range(0, len(open("origin_timetable.txt").readlines())):
    line = data.readline()
    stations = re.findall(r'\{([\s\S]+?)\}', line)
    station_num = len(stations) - 1
    sheet.cell(1, 1, 'No')
    sheet.cell(1, 2, 'TrainNumber')
    sheet.cell(1, 3, 'TrainClass')
    sheet.cell(1, 4, 'Station')
    sheet.cell(1, 5, 'ArriveTime')
    sheet.cell(1, 6, 'DepartureTime')
    sheet.cell(1, 7, 'StayTime')
    sheet.cell(1, 8, 'ArriveDay')
    sheet.cell(1, 9, 'RunTime')
    sheet.cell(1, 10, 'TrainCode')
    sheet.cell(1, 11, 'ID')
    for num in range(0, station_num):
        sheet2.cell(num + line_gap - 1, 1,
                    str(re.findall(r'\"station_train_code\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2])
        sheet2.cell(num + line_gap - 1, 2, str(stations[station_num]))
        sheet.cell(num + line_gap, 1, str(re.findall(r'\"station_no\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2])
        sheet.cell(num + line_gap, 2,
                   str(re.findall(r'\"station_train_code\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2])
        sheet.cell(num + line_gap, 3, str(re.findall(r'\"train_class_name\"\:\"([\s\S]+?)\"', str(stations[0])))[2:-2])
        sheet.cell(num + line_gap, 4, str(re.findall(r'\"station_name\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2])
        arrive_time = str(re.findall(r'\"arrive_time\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2]
        sheet.cell(num + line_gap, 5, arrive_time)
        departure_time = str(re.findall(r'\"start_time\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2]
        sheet.cell(num + line_gap, 6, departure_time)
        if arrive_time == '----' or departure_time == '----':
            stop_over = '----'
        else:
            t1 = datetime.datetime.strptime(arrive_time, '%H:%M')
            t2 = datetime.datetime.strptime(departure_time, '%H:%M')
            if arrive_time < departure_time:
                secondsDiff = (t2 - t1).total_seconds()
                minutesDiff = int(secondsDiff / 60)
                if minutesDiff < 200:
                    stop_over = str(minutesDiff) + '分'
                else:
                    stop_over = '----'
            elif arrive_time > departure_time:
                secondsDiff = (t2 + datetime.timedelta(days=1) - t1).total_seconds()
                minutesDiff = int(secondsDiff / 60)
                if minutesDiff < 200:
                    stop_over = str(minutesDiff) + '分'
                else:
                    stop_over = '----'
            else:
                stop_over = '----'
        sheet.cell(num + line_gap, 7, stop_over)
        sheet.cell(num + line_gap, 8, str(re.findall(r'\"arrive_day_str\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2])
        sheet.cell(num + line_gap, 9, str(re.findall(r'\"running_time\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2])
        sheet.cell(num + line_gap, 10, str(stations[station_num]))
        sheet.cell(num + line_gap, 11,
                   str(stations[station_num]) +
                   str(re.findall(r'\"station_no\"\:\"([\s\S]+?)\"', str(stations[num])))[2:-2])
    line_gap = line_gap + station_num
workbook.save(filename="timetable.xlsx")
codelist.save(filename='codelist.xlsx')
xlsx = pd.read_excel('codelist.xlsx', index_col=0)
xlsx.to_csv('codelist.csv', encoding='utf-8')
csv = pd.read_csv('codelist.csv', low_memory=False, error_bad_lines=False, header=None)  # 读取csv中的数据
df = pd.DataFrame(csv)
f = df.drop_duplicates(subset=0)  # 去重
f.to_csv('no_code.csv', index=None)  # 写到一个新的文件
new_csv = pd.read_csv('no_code.csv', encoding='utf-8', header=0, names=['TrainNumber', 'TrainCode'])
new_csv.to_excel('no_code.xlsx', sheet_name='trains')
