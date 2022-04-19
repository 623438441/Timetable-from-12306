from shutil import copyfile

import openpyxl

copyfile('train.xlsx', 'trainlist.xlsx')

trainlist = openpyxl.load_workbook('trainlist.xlsx')
no_code = openpyxl.load_workbook('no_code.xlsx')
listsht = trainlist.get_sheet_by_name('trains')
codesht = no_code.get_sheet_by_name('trains')
for row in range(2, codesht.max_row + 1):
    code = codesht.cell(row=row, column=3)
    for row2 in range(2, listsht.max_row + 1):
        code2 = listsht.cell(row=row2, column=3)
        if str(code.value) != str(code2.value):
            continue
        else:
            if (listsht.cell(row=row2, column=2)).value is None:
                listsht.cell(row2, 2, str((codesht.cell(row=row, column=2)).value))
            else:
                listsht.cell(row2, 2, str((listsht.cell(row=row2, column=2)).value) + '/' + str(
                    (codesht.cell(row=row, column=2)).value))
trainlist.save('trainlist.xlsx')
