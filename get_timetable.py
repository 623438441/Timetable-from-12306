import re

import openpyxl
import requests

xlsx = openpyxl.load_workbook('train.xlsx')
sht = xlsx['trains']
file = open('origin_timetable.txt', 'w')
session = requests.session()
date = input('YYYY-MM-DD')
for lines in range(2, sht.max_row + 1):
    train_no = str(sht.cell(lines, 3).value)
    html = session.get(
        f'https://kyfw.12306.cn/otn/queryTrainInfo/query?leftTicketDTO.train_no={train_no}&leftTicketDTO.train_date={date}&rand_code=')
    text = re.sub(r'\{\"validateMessagesShowId([\s\S]+?)\:\[', '', html.text)
    text = re.sub(r'\]\}\,\"messages([\s\S]+?)validateMessages\"\:\{\}\}', '', text)
    file.write(text + ',{' + f'{train_no}' + '}\n')
file.close()
